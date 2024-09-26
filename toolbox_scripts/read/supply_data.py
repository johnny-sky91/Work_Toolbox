import pandas as pd
import datetime, os, openpyxl

from datetime import timedelta
from openpyxl.worksheet.datavalidation import DataValidation


class SupplyDataReader:
    def __init__(self, path_supply: str = None, path_groups: str = None):
        self.path_supply = path_supply
        self.path_groups = path_groups
        self.data = None
        self.groups = None
        self.supply_confirmed = None
        self.supply_requested = None

    def read_data(self):
        self.data = pd.read_excel(self.path_supply, sheet_name="SUPPLY")

    def read_groups(self):
        self.groups = pd.read_excel(self.path_groups, sheet_name="groups")

    def prepare_data(self):
        self.data = self.data.iloc[1:, 4:]
        self.data = self.data.reset_index(drop=True)
        headers_row = 0
        new_headers = self.data.iloc[headers_row].tolist()
        self.data = self.data.rename(columns=dict(zip(self.data.columns, new_headers)))
        self.data = self.data.drop(self.data.index[0])
        self.data = self.data.rename(columns={self.data.columns[0]: "COMPONENT"})

    def select_data(self):
        ready_columns = ["COMPONENT", "factory\n(destination)", "date", "Supplier"]
        date_columns = [
            x for x in self.data.columns if isinstance(x, datetime.datetime)
        ]
        today = datetime.datetime.today()
        today = today.replace(hour=0, minute=0, second=0, microsecond=0)
        start_of_week = today - timedelta(days=today.weekday())

        date_columns = date_columns[date_columns.index(start_of_week) :]
        self.data = self.data[ready_columns + date_columns]

    def filter_data(self, filtered, searched):
        filtered = pd.DataFrame(
            self.data[
                (self.data["date"].str.contains(searched))
                & (self.data["factory\n(destination)"] == "CZ")
            ]
        )
        filtered.rename(columns={"Supplier": "SUPPLIER"}, inplace=True)
        filtered.drop(columns=["date", "factory\n(destination)"], inplace=True)
        filtered.infer_objects(copy=False).fillna(0)
        return filtered

    def add_groups(self, data):
        self.groups = self.groups[["COMPONENT", "GROUP"]].drop_duplicates()
        self.groups.reset_index(inplace=True, drop=True)
        data = pd.merge(data, self.groups, on="COMPONENT", how="inner")
        data.insert(1, "GROUP", data.pop("GROUP"))
        return data

    def melt_supply(self, data):
        melted_data = data.melt(
            id_vars=["GROUP", "COMPONENT", "SUPPLIER"],
            var_name="ETD_DATE_WEEK",
            value_name="QTY",
        )
        melted_data = melted_data.loc[melted_data["QTY"] != 0]
        melted_data = melted_data.dropna(subset=["QTY"])
        melted_data.reset_index(inplace=True, drop=True)
        melted_data[["STATUS", "SHIPMENT_ID", "COMMENT", "ETA_DATE_WEEK", "IN_SAP"]] = (
            None
        )
        return melted_data

    def save_to_excel(self):
        current_datetime = datetime.datetime.now()

        current_year = current_datetime.strftime("%Y")
        current_week = current_datetime.strftime("%V")
        filename = f"new_Components_supply_CW{current_week}_{current_year}.xlsx"
        directory_path = os.path.dirname(self.path_supply)
        report_file_path = os.path.join(directory_path, filename)
        source_file = os.path.basename(self.path_supply)

        writer = pd.ExcelWriter(report_file_path)

        self.supply_confirmed.to_excel(
            writer, sheet_name=f"supply_confirmed", index=False
        )
        self.supply_requested.to_excel(
            writer, sheet_name=f"supply_requested", index=False
        )

        info_df = pd.DataFrame({"Source_file": [source_file]})
        info_df.to_excel(writer, sheet_name="INFO", index=False)

        writer._save()
        add_dropdown_statuses(
            excel_filename=report_file_path, sheet_name="supply_confirmed"
        )

    def __call__(self):
        self.read_data()
        self.read_groups()
        self.prepare_data()
        self.select_data()

        self.supply_confirmed = self.filter_data(
            filtered=self.supply_confirmed, searched=": C"
        )
        self.supply_requested = self.filter_data(
            filtered=self.supply_requested, searched=":B"
        )

        self.supply_confirmed = self.add_groups(self.supply_confirmed)
        self.supply_requested = self.add_groups(self.supply_requested)

        self.supply_confirmed = self.melt_supply(self.supply_confirmed)
        self.supply_requested = self.melt_supply(self.supply_requested)

        self.save_to_excel()


def add_dropdown_statuses(excel_filename: str, sheet_name: str):
    workbook = openpyxl.load_workbook(excel_filename)
    sheet = workbook[sheet_name]

    status = ["Shipped", "Open_WH", "Delivered", "Other"]
    status_ready = '"' + ",".join(status) + '"'

    status_validation = DataValidation(
        type="list", formula1=status_ready, allow_blank=True
    )

    sheet.add_data_validation(status_validation)
    status_validation_range = "F2:F" + str(sheet.max_row)
    status_validation.add(status_validation_range)

    in_sap = ["True", "False"]
    in_sap_ready = '"' + ",".join(in_sap) + '"'

    in_sap_validation = DataValidation(
        type="list", formula1=in_sap_ready, allow_blank=True
    )

    sheet.add_data_validation(in_sap_validation)
    in_sap_validation_range = "J2:J" + str(sheet.max_row)
    in_sap_validation.add(in_sap_validation_range)

    workbook.save(excel_filename)
